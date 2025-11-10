from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import TipoLote, TipoLotePrecoEstrela, TipoLotePrecoClassificacao, db, FornecedorTipoLoteClassificacao, Fornecedor
from app.auth import admin_required
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

bp = Blueprint('tipos_lote', __name__, url_prefix='/api/tipos-lote')

@bp.route('', methods=['GET'])
@jwt_required()
def listar_tipos_lote():
    try:
        busca = request.args.get('busca', '')
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = TipoLote.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if busca:
            query = query.filter(
                db.or_(
                    TipoLote.nome.ilike(f'%{busca}%'),
                    TipoLote.codigo.ilike(f'%{busca}%'),
                    TipoLote.descricao.ilike(f'%{busca}%')
                )
            )
        
        tipos = query.order_by(TipoLote.nome).all()
        return jsonify([tipo.to_dict() for tipo in tipos]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar tipos de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        return jsonify(tipo.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter tipo de lote: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_tipo_lote():
    try:
        data = request.get_json()
        
        if not data or not data.get('nome'):
            return jsonify({'erro': 'Nome é obrigatório'}), 400
        
        tipo_existente = TipoLote.query.filter_by(nome=data['nome']).first()
        if tipo_existente:
            return jsonify({'erro': 'Já existe um tipo de lote com este nome'}), 400
        
        if data.get('codigo'):
            codigo_existente = TipoLote.query.filter_by(codigo=data['codigo']).first()
            if codigo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este código'}), 400
        
        total_tipos = TipoLote.query.count()
        if total_tipos >= 150:
            return jsonify({'erro': 'Limite máximo de 150 tipos de lote atingido'}), 400
        
        classificacao = data.get('classificacao', None)
        if classificacao and classificacao not in ['leve', 'media', 'pesada', '']:
            return jsonify({'erro': 'Classificação deve ser: leve, media ou pesada'}), 400
        
        if classificacao == '':
            classificacao = None
        
        tipo = TipoLote(
            nome=data['nome'],
            descricao=data.get('descricao', ''),
            codigo=data.get('codigo', ''),
            classificacao=classificacao,
            ativo=data.get('ativo', True)
        )
        
        db.session.add(tipo)
        db.session.commit()
        
        return jsonify(tipo.to_dict()), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar tipo de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        data = request.get_json()
        
        if not data:
            return jsonify({'erro': 'Dados não fornecidos'}), 400
        
        if data.get('nome') and data['nome'] != tipo.nome:
            tipo_existente = TipoLote.query.filter_by(nome=data['nome']).first()
            if tipo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este nome'}), 400
            tipo.nome = data['nome']
        
        if data.get('codigo') and data['codigo'] != tipo.codigo:
            codigo_existente = TipoLote.query.filter_by(codigo=data['codigo']).first()
            if codigo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este código'}), 400
            tipo.codigo = data['codigo']
        
        if 'descricao' in data:
            tipo.descricao = data['descricao']
        
        if 'classificacao' in data:
            classificacao = data['classificacao']
            if classificacao == '':
                classificacao = None
            if classificacao and classificacao not in ['leve', 'media', 'pesada']:
                return jsonify({'erro': 'Classificação deve ser: leve, media ou pesada'}), 400
            tipo.classificacao = classificacao
        
        if 'ativo' in data:
            tipo.ativo = data['ativo']
        
        db.session.commit()
        
        return jsonify(tipo.to_dict()), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar tipo de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        if len(tipo.itens_solicitacao) > 0 or len(tipo.lotes) > 0:
            return jsonify({'erro': 'Não é possível deletar tipo de lote com solicitações ou lotes associados. Desative-o em vez disso.'}), 400
        
        db.session.delete(tipo)
        db.session.commit()
        
        return jsonify({'mensagem': 'Tipo de lote deletado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar tipo de lote: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        arquivo = request.files['arquivo']
        
        if not arquivo.filename or arquivo.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Formato de arquivo inválido. Use .xlsx ou .xls'}), 400
        
        total_tipos = TipoLote.query.count()
        
        arquivo_bytes = arquivo.read()
        df_tipos = pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name='Tipos de Lote') if 'Tipos de Lote' in pd.ExcelFile(io.BytesIO(arquivo_bytes)).sheet_names else pd.read_excel(io.BytesIO(arquivo_bytes))
        
        colunas_requeridas = ['nome']
        colunas_faltando = [col for col in colunas_requeridas if col not in df_tipos.columns]
        
        if colunas_faltando:
            return jsonify({'erro': f'Colunas obrigatórias faltando: {", ".join(colunas_faltando)}'}), 400
        
        tipos_criados = 0
        tipos_atualizados = 0
        estrelas_criadas = 0
        estrelas_atualizadas = 0
        erros = []
        
        for idx, (index, row) in enumerate(df_tipos.iterrows()):
            linha_num = idx + 2
            try:
                nome = str(row['nome']).strip()
                if not nome or nome == 'nan':
                    erros.append(f'Linha {linha_num}: Nome é obrigatório')
                    continue
                
                if total_tipos + tipos_criados >= 150:
                    erros.append(f'Linha {linha_num}: Limite máximo de 150 tipos de lote atingido')
                    break
                
                codigo = str(row.get('codigo', '')).strip()
                if codigo == 'nan':
                    codigo = ''
                
                descricao = str(row.get('descricao', '')).strip()
                if descricao == 'nan':
                    descricao = ''
                
                classificacao = str(row.get('classificacao', 'media')).strip().lower()
                if classificacao == 'nan':
                    classificacao = 'media'
                
                if classificacao not in ['leve', 'media', 'pesada']:
                    erros.append(f'Linha {linha_num}: Classificação inválida "{classificacao}". Use: leve, media ou pesada')
                    continue
                
                tipo_existente = TipoLote.query.filter_by(nome=nome).first()
                
                if tipo_existente:
                    tipo_existente.descricao = descricao if descricao else tipo_existente.descricao
                    tipo_existente.classificacao = classificacao
                    if codigo:
                        codigo_em_uso = TipoLote.query.filter(
                            TipoLote.codigo == codigo,
                            TipoLote.id != tipo_existente.id
                        ).first()
                        if not codigo_em_uso:
                            tipo_existente.codigo = codigo
                    tipos_atualizados += 1
                else:
                    if codigo:
                        codigo_existente = TipoLote.query.filter_by(codigo=codigo).first()
                        if codigo_existente:
                            erros.append(f'Linha {linha_num}: Código "{codigo}" já está em uso')
                            continue
                    
                    novo_tipo = TipoLote(
                        nome=nome,
                        codigo=codigo,
                        descricao=descricao,
                        classificacao=classificacao,
                        ativo=True
                    )
                    db.session.add(novo_tipo)
                    tipos_criados += 1
            
            except Exception as e:
                erros.append(f'Linha {linha_num}: {str(e)}')
                continue
        
        db.session.commit()
        
        try:
            excel_file = pd.ExcelFile(io.BytesIO(arquivo_bytes))
            if 'Estrelas por Fornecedor' in excel_file.sheet_names:
                df_estrelas = pd.read_excel(excel_file, sheet_name='Estrelas por Fornecedor')
                
                if 'Tipo de Lote' in df_estrelas.columns and 'Fornecedor' in df_estrelas.columns:
                    for idx, (_, row) in enumerate(df_estrelas.iterrows()):
                        linha_num = idx + 2
                        try:
                            tipo_nome = str(row['Tipo de Lote']).strip()
                            fornecedor_nome = str(row['Fornecedor']).strip()
                            
                            if tipo_nome == 'nan' or fornecedor_nome == 'nan':
                                continue
                            
                            tipo = TipoLote.query.filter_by(nome=tipo_nome).first()
                            fornecedor = Fornecedor.query.filter_by(nome=fornecedor_nome).first()
                            
                            if not tipo or not fornecedor:
                                erros.append(f'Estrelas L{linha_num}: Tipo ou fornecedor não encontrado')
                                continue
                            
                            leve_val = row.get('Leve (⭐)', 1)
                            leve = int(leve_val) if not pd.isna(leve_val) else 1
                            medio_val = row.get('Médio (⭐)', 3)
                            medio = int(medio_val) if not pd.isna(medio_val) else 3
                            pesado_val = row.get('Pesado (⭐)', 5)
                            pesado = int(pesado_val) if not pd.isna(pesado_val) else 5
                            
                            ativo_val = row.get('Ativo')
                            if pd.isna(ativo_val) or ativo_val == '':
                                ativo = True
                            else:
                                ativo = str(ativo_val).strip().lower() in ['sim', 'yes', '1', 'true']
                            
                            classif_existente = FornecedorTipoLoteClassificacao.query.filter_by(
                                fornecedor_id=fornecedor.id,
                                tipo_lote_id=tipo.id
                            ).first()
                            
                            if classif_existente:
                                classif_existente.leve_estrelas = leve
                                classif_existente.medio_estrelas = medio
                                classif_existente.pesado_estrelas = pesado
                                classif_existente.ativo = ativo
                                estrelas_atualizadas += 1
                            else:
                                nova_classif = FornecedorTipoLoteClassificacao(
                                    fornecedor_id=fornecedor.id,
                                    tipo_lote_id=tipo.id,
                                    leve_estrelas=leve,
                                    medio_estrelas=medio,
                                    pesado_estrelas=pesado,
                                    ativo=ativo
                                )
                                db.session.add(nova_classif)
                                estrelas_criadas += 1
                        except Exception as e:
                            erros.append(f'Estrelas L{linha_num}: {str(e)}')
                            continue
                    
                    db.session.commit()
        except Exception as e:
            erros.append(f'Erro ao processar estrelas: {str(e)}')
        
        return jsonify({
            'mensagem': 'Importação concluída',
            'tipos_criados': tipos_criados,
            'tipos_atualizados': tipos_atualizados,
            'estrelas_criadas': estrelas_criadas,
            'estrelas_atualizadas': estrelas_atualizadas,
            'erros': erros
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar arquivo: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@admin_required
def exportar_excel():
    try:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        ws_tipos = wb.create_sheet('Tipos de Lote')
        if default_sheet and default_sheet in wb.worksheets:
            wb.remove(default_sheet)
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        headers_tipos = ['ID', 'Código', 'Nome', 'Classificação', 'Descrição', 'Ativo']
        ws_tipos.append(headers_tipos)
        
        for cell in ws_tipos[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        tipos = TipoLote.query.order_by(TipoLote.nome).all()
        for tipo in tipos:
            ws_tipos.append([
                tipo.id,
                tipo.codigo or '',
                tipo.nome,
                tipo.classificacao,
                tipo.descricao or '',
                'Sim' if tipo.ativo else 'Não'
            ])
        
        for column in ws_tipos.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_tipos.column_dimensions[column[0].column_letter].width = adjusted_width
        
        ws_estrelas = wb.create_sheet('Estrelas por Fornecedor')
        headers_estrelas = ['Tipo de Lote', 'Fornecedor', 'Leve (⭐)', 'Médio (⭐)', 'Pesado (⭐)', 'Ativo']
        ws_estrelas.append(headers_estrelas)
        
        for cell in ws_estrelas[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        classificacoes = FornecedorTipoLoteClassificacao.query.join(
            TipoLote, FornecedorTipoLoteClassificacao.tipo_lote_id == TipoLote.id
        ).join(
            Fornecedor, FornecedorTipoLoteClassificacao.fornecedor_id == Fornecedor.id
        ).order_by(TipoLote.nome, Fornecedor.nome).all()
        
        for classif in classificacoes:
            ws_estrelas.append([
                classif.tipo_lote.nome if classif.tipo_lote else '',
                classif.fornecedor.nome if classif.fornecedor else '',
                classif.leve_estrelas,
                classif.medio_estrelas,
                classif.pesado_estrelas,
                'Sim' if classif.ativo else 'Não'
            ])
        
        for column in ws_estrelas.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_estrelas.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'tipos_lote_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar Excel: {str(e)}'}), 500

@bp.route('/<int:tipo_id>/precos', methods=['GET'])
@jwt_required()
def listar_precos_estrela(tipo_id):
    try:
        tipo = TipoLote.query.get(tipo_id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        precos = TipoLotePrecoEstrela.query.filter_by(tipo_lote_id=tipo_id).order_by(TipoLotePrecoEstrela.estrelas).all()
        
        return jsonify({
            'tipo_lote': tipo.to_dict(),
            'precos': [preco.to_dict() for preco in precos]
        }), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar preços: {str(e)}'}), 500

@bp.route('/<int:tipo_id>/precos', methods=['POST'])
@admin_required
def configurar_precos_estrela(tipo_id):
    try:
        tipo = TipoLote.query.get(tipo_id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        data = request.get_json()
        
        if not data or 'precos' not in data:
            return jsonify({'erro': 'Lista de preços não fornecida'}), 400
        
        precos_inseridos = 0
        precos_atualizados = 0
        
        for preco_data in data['precos']:
            estrelas = preco_data.get('estrelas')
            preco_kg = preco_data.get('preco_por_kg', 0.0)
            
            if not estrelas or not (1 <= estrelas <= 5):
                continue
            
            preco_existente = TipoLotePrecoEstrela.query.filter_by(
                tipo_lote_id=tipo_id,
                estrelas=estrelas
            ).first()
            
            if preco_existente:
                preco_existente.preco_por_kg = preco_kg
                preco_existente.ativo = preco_data.get('ativo', True)
                precos_atualizados += 1
            else:
                novo_preco = TipoLotePrecoEstrela(
                    tipo_lote_id=tipo_id,
                    estrelas=estrelas,
                    preco_por_kg=preco_kg,
                    ativo=preco_data.get('ativo', True)
                )
                db.session.add(novo_preco)
                precos_inseridos += 1
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Preços configurados com sucesso',
            'inseridos': precos_inseridos,
            'atualizados': precos_atualizados
        }), 200
    
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao configurar preços: {str(e)}'}), 500

@bp.route('/<int:tipo_id>/precos/<int:estrelas>', methods=['DELETE'])
@admin_required
def deletar_preco_estrela(tipo_id, estrelas):
    try:
        if not (1 <= estrelas <= 5):
            return jsonify({'erro': 'Estrelas deve estar entre 1 e 5'}), 400
        
        preco = TipoLotePrecoEstrela.query.filter_by(
            tipo_lote_id=tipo_id,
            estrelas=estrelas
        ).first()
        
        if not preco:
            return jsonify({'erro': 'Preço não encontrado'}), 404
        
        db.session.delete(preco)
        db.session.commit()
        
        return jsonify({'mensagem': 'Preço deletado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar preço: {str(e)}'}), 500

@bp.route('/<int:tipo_id>/precos-classificacao', methods=['GET'])
@jwt_required()
def listar_precos_classificacao(tipo_id):
    try:
        tipo = TipoLote.query.get(tipo_id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        precos = TipoLotePrecoClassificacao.query.filter_by(tipo_lote_id=tipo_id).all()
        
        return jsonify({
            'tipo_lote': tipo.to_dict(),
            'precos': [preco.to_dict() for preco in precos]
        }), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar preços de classificação: {str(e)}'}), 500

@bp.route('/<int:tipo_id>/precos-classificacao', methods=['POST'])
@admin_required
def configurar_precos_classificacao(tipo_id):
    try:
        tipo = TipoLote.query.get(tipo_id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        data = request.get_json()
        
        if not data:
            return jsonify({'erro': 'Dados não fornecidos'}), 400
        
        precos_inseridos = 0
        precos_atualizados = 0
        
        for classificacao in ['leve', 'medio', 'pesado']:
            preco_key = f'{classificacao}_preco'
            
            if preco_key not in data:
                continue
            
            preco_kg = data[preco_key]
            
            if preco_kg is None:
                continue
                
            if preco_kg < 0:
                return jsonify({'erro': f'Preço para {classificacao} não pode ser negativo'}), 400
            
            preco_existente = TipoLotePrecoClassificacao.query.filter_by(
                tipo_lote_id=tipo_id,
                classificacao=classificacao
            ).first()
            
            if preco_existente:
                preco_existente.preco_por_kg = preco_kg
                preco_existente.ativo = data.get('ativo', True)
                precos_atualizados += 1
            else:
                novo_preco = TipoLotePrecoClassificacao(
                    tipo_lote_id=tipo_id,
                    classificacao=classificacao,
                    preco_por_kg=preco_kg,
                    ativo=data.get('ativo', True)
                )
                db.session.add(novo_preco)
                precos_inseridos += 1
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Preços de classificação configurados com sucesso',
            'inseridos': precos_inseridos,
            'atualizados': precos_atualizados
        }), 200
    
    except ValueError as e:
        db.session.rollback()
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao configurar preços de classificação: {str(e)}'}), 500
