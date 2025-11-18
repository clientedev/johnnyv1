from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models import db, MaterialBase, TabelaPreco, TabelaPrecoItem, Usuario
from app.auth import admin_required
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

bp = Blueprint('materiais_base', __name__, url_prefix='/api/materiais-base')

def gerar_codigo_automatico():
    ultimo_material = MaterialBase.query.order_by(MaterialBase.id.desc()).first()
    if ultimo_material and ultimo_material.codigo:
        try:
            numero = int(ultimo_material.codigo.replace('MAT', ''))
            return f'MAT{numero + 1:03d}'
        except:
            pass
    
    proximo_id = MaterialBase.query.count() + 1
    return f'MAT{proximo_id:03d}'

@bp.route('', methods=['GET'])
@jwt_required()
def listar_materiais():
    try:
        busca = request.args.get('busca', '')
        classificacao = request.args.get('classificacao', '')
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = MaterialBase.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if busca:
            query = query.filter(
                db.or_(
                    MaterialBase.nome.ilike(f'%{busca}%'),
                    MaterialBase.codigo.ilike(f'%{busca}%')
                )
            )
        
        if classificacao and classificacao in ['leve', 'medio', 'pesado']:
            query = query.filter_by(classificacao=classificacao)
        
        materiais = query.order_by(MaterialBase.codigo).all()
        
        result = []
        for material in materiais:
            material_dict = material.to_dict()
            result.append(material_dict)
        
        return jsonify(result), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar materiais: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_material(id):
    try:
        material = MaterialBase.query.get(id)
        
        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404
        
        return jsonify(material.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter material: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_material():
    try:
        data = request.get_json()
        
        print("=" * 80)
        print("DEBUG: CRIAR MATERIAL")
        print(f"Dados completos recebidos: {data}")
        print(f"Tipo de 'precos': {type(data.get('precos'))}")
        print(f"Conteúdo de 'precos': {data.get('precos')}")
        if data.get('precos'):
            for key, value in data.get('precos').items():
                print(f"  Chave: {key}, Valor: {value}, Tipo: {type(value)}")
        print("=" * 80)
        
        required_fields = ['nome', 'classificacao']
        for field in required_fields:
            if field not in data:
                return jsonify({'erro': f'Campo {field} é obrigatório'}), 400
        
        if data['classificacao'] not in ['leve', 'medio', 'pesado']:
            return jsonify({'erro': 'Classificação deve ser: leve, medio ou pesado'}), 400
        
        material_existente = MaterialBase.query.filter_by(nome=data['nome']).first()
        if material_existente:
            return jsonify({'erro': 'Já existe um material com este nome'}), 400
        
        codigo = data.get('codigo') or gerar_codigo_automatico()
        codigo_existente = MaterialBase.query.filter_by(codigo=codigo).first()
        if codigo_existente:
            return jsonify({'erro': 'Código já está em uso'}), 400
        
        total_materiais = MaterialBase.query.count()
        if total_materiais >= 500:
            return jsonify({'erro': 'Limite de 500 materiais atingido'}), 400
        
        material = MaterialBase(
            codigo=codigo,
            nome=data['nome'],
            classificacao=data['classificacao'],
            descricao=data.get('descricao', ''),
            ativo=True
        )
        db.session.add(material)
        db.session.flush()
        
        tabelas_preco = TabelaPreco.query.all()
        precos = data.get('precos', {})
        
        print(f"DEBUG: Número de tabelas encontradas: {len(tabelas_preco)}")
        for tabela in tabelas_preco:
            print(f"DEBUG: Processando tabela ID={tabela.id}, nivel_estrelas={tabela.nivel_estrelas}")
            preco_key = f'preco_{tabela.nivel_estrelas}_estrela'
            preco_valor_raw = precos.get(preco_key)
            
            print(f"DEBUG: Buscando chave '{preco_key}' nos precos")
            print(f"DEBUG: Valor raw encontrado: {preco_valor_raw} (tipo: {type(preco_valor_raw)})")
            
            try:
                if preco_valor_raw is None or preco_valor_raw == '':
                    preco_valor = 0.00
                    print(f"DEBUG: Valor vazio/None, usando 0.00")
                else:
                    preco_valor = float(preco_valor_raw)
                    print(f"DEBUG: Convertido para float: {preco_valor}")
            except (ValueError, TypeError) as e:
                preco_valor = 0.00
                print(f"DEBUG: Erro na conversão ({e}), usando 0.00")
            
            print(f"DEBUG: Salvando preco_por_kg={preco_valor} para tabela {tabela.nivel_estrelas} estrelas")
            
            preco_item = TabelaPrecoItem(
                tabela_preco_id=tabela.id,
                material_id=material.id,
                preco_por_kg=preco_valor,
                ativo=True
            )
            db.session.add(preco_item)
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Material criado com sucesso',
            'material': material.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar material: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_material(id):
    try:
        material = MaterialBase.query.get(id)
        
        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404
        
        data = request.get_json()
        
        if 'nome' in data and data['nome'] != material.nome:
            material_existente = MaterialBase.query.filter_by(nome=data['nome']).first()
            if material_existente:
                return jsonify({'erro': 'Já existe um material com este nome'}), 400
            material.nome = data['nome']
        
        if 'classificacao' in data:
            if data['classificacao'] not in ['leve', 'medio', 'pesado']:
                return jsonify({'erro': 'Classificação deve ser: leve, medio ou pesado'}), 400
            material.classificacao = data['classificacao']
        
        if 'descricao' in data:
            material.descricao = data['descricao']
        
        if 'ativo' in data:
            material.ativo = data['ativo']
        
        if 'precos' in data:
            precos = data['precos']
            for preco_item in material.precos:
                tabela = preco_item.tabela_preco
                if tabela:
                    preco_key = f'preco_{tabela.nivel_estrelas}_estrela'
                    if preco_key in precos:
                        preco_item.preco_por_kg = float(precos[preco_key])
        
        material.data_atualizacao = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Material atualizado com sucesso',
            'material': material.to_dict()
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar material: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_material(id):
    try:
        material = MaterialBase.query.get(id)
        
        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404
        
        material.ativo = False
        db.session.commit()
        
        return jsonify({'mensagem': 'Material desativado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar material: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        
        df = pd.read_excel(file)
        
        colunas_esperadas = ['Nome do Material', 'Classificação', '1 Estrela (R$/kg)', '2 Estrelas (R$/kg)', '3 Estrelas (R$/kg)']
        colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]
        
        if colunas_faltando:
            return jsonify({'erro': f'Colunas faltando no Excel: {", ".join(colunas_faltando)}'}), 400
        
        tabelas_preco = {tab.nivel_estrelas: tab for tab in TabelaPreco.query.all()}
        
        if len(tabelas_preco) < 3:
            return jsonify({'erro': 'É necessário ter as 3 tabelas de preço cadastradas'}), 400
        
        materiais_criados = 0
        materiais_atualizados = 0
        precos_atualizados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                nome = str(row['Nome do Material']).strip()
                classificacao = str(row['Classificação']).strip().lower()
                
                if classificacao not in ['leve', 'medio', 'pesado']:
                    erros.append(f"Linha {index+2}: Classificação inválida '{classificacao}'")
                    continue
                
                material = MaterialBase.query.filter_by(nome=nome).first()
                
                if not material:
                    codigo = gerar_codigo_automatico()
                    material = MaterialBase(
                        codigo=codigo,
                        nome=nome,
                        classificacao=classificacao,
                        ativo=True
                    )
                    db.session.add(material)
                    db.session.flush()
                    
                    for nivel_estrelas, tabela in tabelas_preco.items():
                        preco_col = f'{nivel_estrelas} {"Estrelas" if nivel_estrelas > 1 else "Estrela"} (R$/kg)'
                        preco_valor = float(row.get(preco_col, 0.00))
                        
                        preco_item = TabelaPrecoItem(
                            tabela_preco_id=tabela.id,
                            material_id=material.id,
                            preco_por_kg=preco_valor,
                            ativo=True
                        )
                        db.session.add(preco_item)
                        precos_atualizados += 1
                    
                    materiais_criados += 1
                else:
                    material.classificacao = classificacao
                    
                    for nivel_estrelas, tabela in tabelas_preco.items():
                        preco_col = f'{nivel_estrelas} {"Estrelas" if nivel_estrelas > 1 else "Estrela"} (R$/kg)'
                        preco_valor = float(row.get(preco_col, 0.00))
                        
                        preco_item = TabelaPrecoItem.query.filter_by(
                            tabela_preco_id=tabela.id,
                            material_id=material.id
                        ).first()
                        
                        if preco_item:
                            preco_item.preco_por_kg = preco_valor
                        else:
                            preco_item = TabelaPrecoItem(
                                tabela_preco_id=tabela.id,
                                material_id=material.id,
                                preco_por_kg=preco_valor,
                                ativo=True
                            )
                            db.session.add(preco_item)
                        precos_atualizados += 1
                    
                    materiais_atualizados += 1
            
            except Exception as e:
                erros.append(f"Linha {index+2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Importação concluída',
            'materiais_criados': materiais_criados,
            'materiais_atualizados': materiais_atualizados,
            'precos_atualizados': precos_atualizados,
            'erros': erros
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar Excel: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@admin_required
def exportar_excel():
    try:
        materiais = MaterialBase.query.order_by(MaterialBase.codigo).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Materiais e Preços'
        
        headers = ['Código', 'Nome do Material', 'Classificação', '1 Estrela (R$/kg)', '2 Estrelas (R$/kg)', '3 Estrelas (R$/kg)', 'Descrição']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for material in materiais:
            precos_dict = material.to_dict().get('precos', {})
            
            row_data = [
                material.codigo,
                material.nome,
                material.classificacao.capitalize(),
                float(precos_dict.get('preco_1_estrela', 0.00)),
                float(precos_dict.get('preco_2_estrela', 0.00)),
                float(precos_dict.get('preco_3_estrela', 0.00)),
                material.descricao or ''
            ]
            ws.append(row_data)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'materiais_base_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar Excel: {str(e)}'}), 500

@bp.route('/modelo-importacao', methods=['GET'])
@admin_required
def modelo_importacao():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Materiais e Preços'
        
        headers = ['Nome do Material', 'Classificação', '1 Estrela (R$/kg)', '2 Estrelas (R$/kg)', '3 Estrelas (R$/kg)', 'Descrição']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        exemplos = [
            ['SUCATA PROCESSADOR CERÂMICO A', 'pesado', 12.50, 15.00, 18.50, 'Processador cerâmico tipo A'],
            ['SUCATA PLACA MÃE SERVIDOR', 'pesado', 8.30, 10.00, 12.00, 'Placa mãe de servidor'],
            ['SUCATA HD SATA', 'medio', 3.50, 4.00, 5.00, 'HD SATA'],
        ]
        
        for exemplo in exemplos:
            ws.append(exemplo)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='modelo_importacao_materiais.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar modelo: {str(e)}'}), 500
